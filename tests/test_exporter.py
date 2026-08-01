import io
import unittest

from openpyxl import load_workbook

from pjsk_salary.exporter import export_salary_excel
from pjsk_salary.parser import parse_schedule_workbook, read_schedule_workbook
from pjsk_salary.salary import build_default_rates, calculate_salary
from pjsk_salary.schedule import build_daily_grids
from tests.test_parser import make_schedule_workbook


class ExporterTests(unittest.TestCase):
    def test_exports_all_sheets_and_summary_total(self):
        parsed = parse_schedule_workbook(make_schedule_workbook())
        rates = build_default_rates(parsed.records)
        result = calculate_salary(parsed.records, rates)
        grids, _ = build_daily_grids(parsed.records, parsed.blocks)

        workbook = load_workbook(
            io.BytesIO(
                export_salary_excel(
                    result,
                    rates,
                    grids,
                )
            ),
            data_only=True,
        )

        self.assertEqual(
            workbook.sheetnames,
            ["工资汇总", "结算明细", "工价设置", "排班-类五"],
        )
        summary = workbook["工资汇总"]
        headers = [cell.value for cell in summary[1]]
        total_row = summary.max_row
        self.assertEqual(summary.cell(total_row, headers.index("班组") + 1).value, "全部")
        self.assertEqual(summary.cell(total_row, headers.index("姓名") + 1).value, "合计")
        self.assertAlmostEqual(
            summary.cell(total_row, headers.index("应发工资") + 1).value,
            result.total_salary,
        )
        schedule = workbook["排班-类五"]
        headers = [cell.value for cell in schedule[2]]
        role_column = headers.index("跑1") + 1
        slot_row = next(
            row
            for row in range(3, schedule.max_row + 1)
            if schedule.cell(row, 1).value == "15:00-15:30"
        )
        self.assertEqual(schedule.cell(slot_row, role_column).value, "九九")
        midnight_row = next(
            row
            for row in range(3, schedule.max_row + 1)
            if schedule.cell(row, 1).value == "00:00-00:30"
        )
        self.assertEqual(schedule.cell(midnight_row, 1).fill.fgColor.rgb, "00F1F3F5")

    def test_escapes_spreadsheet_formulas_in_names_and_headers(self):
        parsed = parse_schedule_workbook(make_schedule_workbook())
        records = parsed.records.iloc[[0]].copy()
        records.loc[:, "role"] = "=危险岗位"
        records.loc[:, "person"] = "@危险姓名"
        rates = build_default_rates(records)
        result = calculate_salary(records, rates)

        workbook = load_workbook(
            io.BytesIO(export_salary_excel(result, rates)),
            read_only=True,
            data_only=False,
        )
        summary = workbook["工资汇总"]

        self.assertEqual(summary["B2"].value, "'@危险姓名")
        self.assertTrue(all(cell.data_type != "f" for row in summary for cell in row))
        self.assertEqual(workbook["工价设置"]["B2"].value, "'=危险岗位")


if __name__ == "__main__":
    unittest.main()
