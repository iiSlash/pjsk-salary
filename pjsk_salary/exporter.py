from __future__ import annotations

import io
import datetime as dt
import re
from typing import Mapping

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .salary import SalaryResult
from .schedule import START_MINUTES_COLUMN, TIME_COLUMN


HEADER_FILL = PatternFill("solid", fgColor="35665C")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="E3EFEB")
TOTAL_BORDER = Border(top=Side(style="thin", color="35665C"))
DATE_FILL = PatternFill("solid", fgColor="EAF2F8")
ROLE_FILL = PatternFill("solid", fgColor="F7E6F5")
ASSIGNMENT_FILL = PatternFill("solid", fgColor="EAF5EE")
NIGHT_FILL = PatternFill("solid", fgColor="F1F3F5")
DATE_TEXT_RE = re.compile(r"^\d{4}-\d{1,2}-\d{1,2}$|^\d{1,2}月\d{1,2}日$")
TIME_TEXT_RE = re.compile(r"^\d{1,2}:\d{2}\s*[-~～—–至]\s*\d{1,2}:\d{2}$")


def export_salary_excel(
    result: SalaryResult,
    rates: pd.DataFrame,
    schedule_grids: Mapping[
        str, Mapping[dt.date, pd.DataFrame] | pd.DataFrame
    ]
    | None = None,
    day_start_hour: int = 8,
    night_start_hour: int = 20,
) -> bytes:
    output = io.BytesIO()
    summary = _summary_with_total(result)
    detail = _safe_frame(result.daily_detail)
    rate_export = _safe_frame(rates)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="工资汇总", index=False)
        detail.to_excel(writer, sheet_name="结算明细", index=False)
        rate_export.to_excel(writer, sheet_name="工价设置", index=False)

        used_sheet_names = set(writer.sheets)
        for team, date_grids in (schedule_grids or {}).items():
            sheet_name = _schedule_sheet_name(str(team), used_sheet_names)
            if isinstance(date_grids, pd.DataFrame):
                _safe_frame(date_grids).to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    header=False,
                )
            else:
                worksheet = writer.book.create_sheet(sheet_name)
                writer.sheets[sheet_name] = worksheet
                _write_daily_schedule(
                    worksheet,
                    date_grids,
                    day_start_hour=day_start_hour,
                    night_start_hour=night_start_hour,
                )
            used_sheet_names.add(sheet_name)

        for sheet_name, worksheet in writer.sheets.items():
            if sheet_name.startswith("排班-"):
                if not any(
                    isinstance(value, Mapping)
                    for value in (schedule_grids or {}).values()
                ):
                    _format_schedule_sheet(worksheet)
                continue

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_cells in worksheet.columns:
                max_length = max(
                    (len(str(cell.value)) if cell.value is not None else 0)
                    for cell in column_cells
                )
                column_letter = column_cells[0].column_letter
                worksheet.column_dimensions[column_letter].width = min(
                    max(max_length + 2, 10), 28
                )

            if sheet_name == "工资汇总":
                for cell in worksheet[worksheet.max_row]:
                    cell.fill = TOTAL_FILL
                    cell.font = Font(bold=True)
                    cell.border = TOTAL_BORDER

            if sheet_name in {"工资汇总", "结算明细", "工价设置"}:
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "#,##0.00"
            for header_cell in worksheet[1]:
                if header_cell.value == "日期":
                    for cell in worksheet.iter_cols(
                        min_col=header_cell.column,
                        max_col=header_cell.column,
                        min_row=2,
                    ):
                        for date_cell in cell:
                            if hasattr(date_cell.value, "year"):
                                date_cell.number_format = "yyyy-mm-dd"

    return output.getvalue()


def _write_daily_schedule(
    worksheet,
    date_grids: Mapping[dt.date, pd.DataFrame],
    *,
    day_start_hour: int,
    night_start_hour: int,
) -> None:
    worksheet.freeze_panes = "B3"
    worksheet.sheet_view.showGridLines = False
    current_row = 1
    maximum_columns = 1

    for schedule_date, frame in sorted(date_grids.items()):
        role_columns = [
            str(column)
            for column in frame.columns
            if column not in {TIME_COLUMN, START_MINUTES_COLUMN}
        ]
        headers = [TIME_COLUMN, *role_columns]
        maximum_columns = max(maximum_columns, len(headers))

        if len(headers) > 1:
            worksheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=len(headers),
            )
        date_cell = worksheet.cell(current_row, 1, schedule_date)
        date_cell.number_format = "yyyy-mm-dd"
        date_cell.fill = DATE_FILL
        date_cell.font = Font(bold=True, color="1F2937")
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[current_row].height = 24
        current_row += 1

        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(current_row, column_index, _safe_text(header))
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[current_row].height = 22
        current_row += 1

        for _, grid_row in frame.sort_values(START_MINUTES_COLUMN, kind="stable").iterrows():
            start_minutes = int(grid_row[START_MINUTES_COLUMN])
            is_night = _is_night_minutes(
                start_minutes,
                day_start_hour=day_start_hour,
                night_start_hour=night_start_hour,
            )
            values = [grid_row[TIME_COLUMN], *[grid_row[role] for role in role_columns]]
            for column_index, value in enumerate(values, start=1):
                cell = worksheet.cell(current_row, column_index, _safe_text(value))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if is_night:
                    cell.fill = NIGHT_FILL
            worksheet.row_dimensions[current_row].height = 20
            current_row += 1
        current_row += 1

    worksheet.column_dimensions["A"].width = 16
    for column_index in range(2, maximum_columns + 1):
        column_letter = get_column_letter(column_index)
        max_length = 10
        for row in worksheet.iter_rows(
            min_col=column_index,
            max_col=column_index,
            max_row=worksheet.max_row,
        ):
            cell = row[0]
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)) + 2)
        worksheet.column_dimensions[column_letter].width = min(max_length, 20)


def _is_night_minutes(
    minutes: int,
    *,
    day_start_hour: int,
    night_start_hour: int,
) -> bool:
    day_start = day_start_hour * 60
    night_start = night_start_hour * 60
    if day_start < night_start:
        return not day_start <= minutes < night_start
    return night_start <= minutes < day_start


def export_summary_csv(result: SalaryResult) -> bytes:
    return _summary_with_total(result).to_csv(index=False).encode("utf-8-sig")


def _summary_with_total(result: SalaryResult) -> pd.DataFrame:
    summary = result.payroll_summary.copy()
    numeric_columns = summary.select_dtypes(include="number").columns
    total_row: dict[str, object] = {column: "" for column in summary.columns}
    total_row["班组"] = "全部"
    total_row["姓名"] = "合计"
    for column in numeric_columns:
        total_row[column] = round(float(summary[column].sum()), 2)
    summary = pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)
    return _safe_frame(summary)


def _schedule_sheet_name(team: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]+", "_", team).strip() or "班组"
    base = f"排班-{cleaned}"[:31]
    candidate = base
    counter = 2
    while candidate in used_names:
        suffix = f"-{counter}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        counter += 1
    return candidate


def _format_schedule_sheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in column_cells
        )
        column_letter = column_cells[0].column_letter
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 10), 18
        )
        for cell in column_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows():
        for position, cell in enumerate(row):
            value = cell.value
            is_date = hasattr(value, "year") or (
                isinstance(value, str) and DATE_TEXT_RE.match(value.strip())
            )
            if is_date:
                cell.fill = DATE_FILL
                cell.font = Font(bold=True)
                if hasattr(value, "year"):
                    cell.number_format = "yyyy-mm-dd"
                _fill_following_cells(row, position, ROLE_FILL, bold=True)
                continue

            if isinstance(value, str) and TIME_TEXT_RE.match(value.strip()):
                _fill_following_cells(row, position, ASSIGNMENT_FILL)


def _fill_following_cells(
    row: tuple,
    start_position: int,
    fill: PatternFill,
    bold: bool = False,
) -> None:
    for following_cell in row[start_position + 1 :]:
        if following_cell.value in (None, ""):
            break
        following_cell.fill = fill
        if bold:
            following_cell.font = Font(bold=True)


def _safe_frame(frame: pd.DataFrame) -> pd.DataFrame:
    safe = frame.copy()
    safe.columns = [_safe_text(column) for column in safe.columns]
    for column in safe.select_dtypes(include=["object", "string"]).columns:
        safe[column] = safe[column].map(_safe_text)
    return safe


def _safe_text(value: object) -> object:
    if not isinstance(value, str) or not value:
        return value
    if value[0] in "=+-@":
        return "'" + value
    return value
